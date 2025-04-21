# Excel_writer
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

import shutil
from tkinter import filedialog, messagebox


EXCEL_PATH = r"resource\Herramienta de Envío de Datos 607.xlsx"
EXCEL_SHEET = "DATOS"

COLUMN_MAP = {
    "RNC/Cédula o Pasaporte": "B",  # 1
    "Tipo Identificación": "C",  # 2
    "Número Comprobante Fiscal": "D",  # 3
    "Número Comprobante Fiscal Modificado": "E",  # 4
    "Tipo de Ingreso": "F",  # 5
    "Fecha Comprobante": "G",  # 6
    "Fecha de Retención": "H",  # 7
    "Monto Facturado": "I",  # 8
    "ITBIS Facturado": "J",  # 9
    "ITBIS Retenido por Terceros": "K",  # 10
    "Cheque/ Transferencia/ Depósito": "L",  # 18
}

# Formatos de columnas
COLUMN_FORMATS = {
    "RNC/Cédula o Pasaporte": "Number",
    "Tipo Identificación": "Text",
    "Número Comprobante Fiscal": "Text",
    "Número Comprobante Fiscal Modificado": "Text",
    "Tipo de Ingreso": "Text",
    "Fecha Comprobante": "Date",
    "Fecha de Retención": "Date",
    "Monto Facturado": "Money",
    "ITBIS Facturado": "Money",
    "ITBIS Retenido por Terceros": "Money",
    "Cheque/ Transferencia/ Depósito": "Money",
}


def aplicar_formato_celda(cell, key, value):
    format_type = COLUMN_FORMATS.get(key)

    if format_type == "Text":
        cell.value = value
        cell.number_format = "@"

    elif format_type == "Date":
        try:
            fecha = datetime.strptime(value.strip(), "%Y%m%d")
            cell.value = fecha
            cell.number_format = "YYYYMMDD"
        except:
            cell.value = value

    elif format_type == "Money":
        try:
            cell.value = float(str(value).replace(",", "").strip())
            cell.number_format = "#,##0.00"
        except:
            cell.value = value

    elif format_type == "Number":
        try:
            cell.value = int(str(value).strip())
            cell.number_format = "0"
        except:
            cell.value = value

    else:
        cell.value = value  # Por si acaso


def guardar_en_excel_existente(data):
    try:
        exito = write_data_to_excel(data)
        if not exito:
            return  # No mostrar nada si falló

        global invoice_data
        invoice_data = None

        messagebox.showinfo(
            "Éxito", "Los datos han sido guardados en el Excel existente."
        )

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al guardar los datos:\n{e}")


# Crear y guardar nuevo archivo desde plantilla
def guardar_en_nuevo_excel(data):
    try:
        plantilla_path = r"resource\Plantilla_de_Formato_607.xlsx"

        nuevo_archivo_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Guardar nuevo Excel como...",
            initialfile="Herramienta de Envío de Datos 607",
        )

        if not nuevo_archivo_path:
            messagebox.showinfo("Cancelado", "No se seleccionó ubicación para guardar.")
            return

        shutil.copy(plantilla_path, nuevo_archivo_path)

        workbook = openpyxl.load_workbook(nuevo_archivo_path)
        sheet = workbook["DATOS"]

        row = 4
        while sheet[f"B{row}"].value:
            row += 1

        for key, col in COLUMN_MAP.items():
            value = data.get(key, "")
            cell = sheet[f"{col}{row}"]
            aplicar_formato_celda(cell, key, value)

        workbook.save(nuevo_archivo_path)
        workbook.close()

        global invoice_data
        invoice_data = None  # Limpieza de memoria

        messagebox.showinfo(
            "Éxito", f"Archivo nuevo guardado correctamente:\n{nuevo_archivo_path}"
        )

    except Exception as e:
        messagebox.showerror(
            "Error", f"Ocurrió un error al guardar el nuevo Excel:\n{e}"
        )


# Función interna: escribe en el archivo base
def write_data_to_excel(data):
    try:
        excel_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos de Excel", "*.xlsx")],
        )

        if not excel_path:
            messagebox.showinfo("Cancelado", "No se seleccionó ningún archivo.")
            return False

        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        row = 4
        while sheet[f"B{row}"].value:
            row += 1

        for key, col in COLUMN_MAP.items():
            value = data.get(key, "")
            cell = sheet[f"{col}{row}"]
            aplicar_formato_celda(cell, key, value)

        workbook.save(excel_path)
        print(f"✅ Datos escritos en la fila {row} del archivo Excel.")
        return True

    except PermissionError:
        print(f"❌ Ocurrió un error al guardar los datos")
        messagebox.showerror(
            "Archivo en uso",
            "No se puede guardar en el archivo porque está abierto.\nCierra el archivo de Excel e intenta de nuevo.",
        )
        return False
